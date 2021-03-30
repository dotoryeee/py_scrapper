from openpyxl import load_workbook
import datetime

wb = load_workbook("test.xlsx")
ws = wb.active

for x in range(1, ws.max_row+1):
    for y in range(1, ws.max_column+1):
        print(ws.cell(row=x, column=y).value, end=" ")

ws = wb["test2"]

print()
print("---------------------------------------------------------")

# test2 시트의 전체 데이터 소환
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        print(cell.value, end="\t")
    print("")

print()
print("---------------------------------------------------------")

# 영어 점수가 90점 이상인 사람의 번호 추출
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    if int(row[2].value) >= 90:
        print(f'{row[0].value}번 학생은 영어 90점 이상')

# ws.insert_rows(8, 5) #8번째 행에 5줄을 추가
# ws.delete_rows(8, 3) #8번쨰 행부터 3줄을 삭제
# ws.delete_cols(2) #B열을 삭제

print()
print("---------------------------------------------------------")

# 값 찾아 바꾸기
for row in ws.iter_rows(max_row=4):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"
            print(cell, "변경 완료")

# 데이터 이동
ws.move_range("B3:C53", rows=0, cols=1)

ws2 = wb.create_sheet("sheet2")

ws2["A1"] = datetime.datetime.today()
ws2["A2"] = "=sum(1, 2, 3)"
ws2["A3"] = '=average(6, 2, 1)'
ws2["A4"] = "=sum(A2:A3)"
wb.save("test.xlsx")
wb.close()

print('워크북을 닫고 data_only 모드로 다시 오픈')
# 워크북을 로드할 때 Data_only 옵션을 이용해 수식이 아닌 결과값을 가져올 수 있다
wb = load_workbook('test.xlsx', data_only=True)
ws = wb["sheet2"]

for rows in ws.values:
    for cell in rows:
        print(cell)
