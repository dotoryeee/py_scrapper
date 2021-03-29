from openpyxl import Workbook

wb = Workbook()  # 워크북 생성
ws = wb.active  # 활성화 된 sheet 가져오기
ws.title = "dotoryeee"  # 시트 이름 변경
ws.sheet_properties.tabColor = "f33333"  # 시트 탭 색상 변경

ws1 = wb.create_sheet("test2")  # 두 번째 시트 생성
ws2 = wb.create_sheet("test3", 1)  # 세 번째 시트를 두 번쨰 탭 위치에 생성

sheetCursor = wb["test2"]  # test2 시트 선택
sheetCursor.sheet_properties.tabColor = "aaaaaa"

sheetCursor["A1"] = "test"

print(wb.sheetnames)

wb.save("test.xlsx")
wb.close()
