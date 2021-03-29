# DOCS : https://openpyxl.readthedocs.io/en/stable/
from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string
import random

wb = Workbook()  # 워크북 생성
ws = wb.active  # 활성화 된 sheet 가져오기
ws.title = "dotoryeee"  # 시트 이름 변경
ws.sheet_properties.tabColor = "f33333"  # 시트 탭 색상 변경

ws1 = wb.create_sheet("test2")  # 두 번째 시트 생성
ws2 = wb.create_sheet("test3", 1)  # 세 번째 시트를 두 번쨰 탭 위치에 생성

sheetCursor = wb["test2"]  # test2 시트 선택
sheetCursor.sheet_properties.tabColor = "aaaaaa"

sheetCursor["A1"] = "test"

print(sheetCursor["A1"])
print(sheetCursor["A1"].value)  # 셀 값 출력
print(sheetCursor.cell(row=1, column=1).value)  # == A1
print(sheetCursor["A2"].value)  # 값이 없으면 None 출력
print("---------------------------------------------------------")
sheetCursor.cell(column=2, row=1, value="yeah")  # == B1
print(sheetCursor["B1"].value)
print("---------------------------------------------------------")
# 5 X 5 랜덤값 채우기
for x in range(1, 6):
    for y in range(1, 6):
        ws2.cell(row=x, column=y, value=random.randint(0, 100))

# 순차적으로 채우기
count = 1
for x in range(1, 6):
    for y in range(1, 6):
        ws.cell(row=x, column=y, value=count)
        count += 1


target = wb.copy_worksheet(sheetCursor)
target.title = "copied sheet"  # 시트 복사

sheetCursor.append(["번호", "수학", "영어"])  # 한 줄씩 데이터 넣기
for i in range(1, 51):
    sheetCursor.append([i, random.randint(1, 100), random.randint(1, 100)])

col_B = sheetCursor["B"]
print(col_B)
print("---------------------------------------------------------")


for cell in col_B:
    print(cell.value, end=" ")

print()
print("---------------------------------------------------------")

col_range = sheetCursor["B:C"]
for cols in col_range:
    for cell in cols:
        print(cell.value, end=" ")
    print()

print("---------------------------------------------------------")
print(wb.sheetnames)

# 표에서 제목을 제외한 데이터만 가져오기
row_range = sheetCursor[4:sheetCursor.max_row]
for rows in row_range:
    for cell in rows:
        print(cell.coordinate, end=": ")
        print(cell.value, end=" ")
    print()

print()
print("---------------------------------------------------------")
for rows in row_range:
    for cell in rows:
        xy = coordinate_from_string(cell.coordinate)
        print(xy, end=" ")
    print()


print()
print("---------------------------------------------------------")

for row in sheetCursor.iter_rows():
    print(row[1].value)  # B열의 모든 값 출력

print()
print("---------------------------------------------------------")
# Range 제한 최종
for row in sheetCursor.iter_rows(min_row=3, max_row=9, min_col=1, max_col=3):
    for cell in row:
        print(cell.value, end="\t")
    print("")

wb.save("test.xlsx")
wb.close()
